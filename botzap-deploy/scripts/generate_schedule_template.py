#!/usr/bin/env python3
"""Шаблон Excel для загрузки расписания на неделю в botzap."""
from __future__ import annotations

import sys
from datetime import date, datetime, time, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

WEEKDAYS = [
    "понедельник",
    "вторник",
    "среда",
    "четверг",
    "пятница",
    "суббота",
    "воскресенье",
]
HOURS = range(9, 23)  # 09:00 … 22:00
STATUS_VALUES = [
    "закрыто",
    "открыто онлайн",
    "открыто ментал хелп",
    "открыто онлайн и ментал хелп",
]


def _week_start(base: date | None = None) -> date:
    base = base or date.today()
    return base - timedelta(days=base.weekday())


def build_template(week_start: date | None = None) -> Workbook:
    week_start = week_start or _week_start()
    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    day_fill = PatternFill("solid", fgColor="D9E2F3")
    time_fill = PatternFill("solid", fgColor="F2F2F2")
    closed_fill = PatternFill("solid", fgColor="FFFFFF")

    ws["A1"] = "Время"
    ws["A1"].font = Font(bold=True)
    ws["A1"].fill = hdr_fill
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center")

    for i, day_name in enumerate(WEEKDAYS):
        col = i + 2
        cell_date = ws.cell(row=1, column=col, value=week_start + timedelta(days=i))
        cell_date.number_format = "DD.MM.YYYY"
        cell_date.fill = hdr_fill
        cell_date.font = Font(bold=True, color="FFFFFF")
        cell_date.alignment = Alignment(horizontal="center")

        cell_day = ws.cell(row=2, column=col, value=day_name)
        cell_day.fill = day_fill
        cell_day.font = Font(bold=True)
        cell_day.alignment = Alignment(horizontal="center")

    ws["A2"] = ""
    ws["A2"].fill = day_fill

    first_data_row = 3
    last_data_row = first_data_row + len(HOURS) - 1
    for r_idx, hour in enumerate(HOURS):
        row = first_data_row + r_idx
        t_cell = ws.cell(row=row, column=1, value=time(hour, 0))
        t_cell.number_format = "HH:MM"
        t_cell.fill = time_fill
        t_cell.font = Font(bold=True)
        t_cell.alignment = Alignment(horizontal="center")
        for col in range(2, 9):
            c = ws.cell(row=row, column=col, value="закрыто")
            c.fill = closed_fill
            c.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=1, max_row=last_data_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = border

    ws.column_dimensions["A"].width = 10
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 22

    ws.freeze_panes = "B3"

    help_ws = wb.create_sheet("Справка")
    help_ws["A1"] = "Как заполнить и отправить в botzap"
    help_ws["A1"].font = Font(bold=True, size=13)
    lines = [
        "",
        "1. На листе «Расписание» отметьте каждый час (09:00–22:00) для каждого дня.",
        "2. Допустимые значения в ячейках (можно выбрать из списка):",
        "   • закрыто",
        "   • открыто онлайн",
        "   • открыто ментал хелп",
        "   • открыто онлайн и ментал хелп",
        "3. Строка с датами — только для вас, бот читает названия дней недели.",
        "4. Сохраните файл .xlsx и отправьте боту: Админ → Расписание → Загрузить Excel.",
        "5. Отправляйте как документ (скрепка → файл), не как фото.",
        "6. Должен быть хотя бы один открытый слот, иначе бот отклонит файл.",
        "",
        "Формат записи клиентов:",
        "   • открыто ментал хелп — очный приём",
        "   • открыто онлайн — онлайн-сессия",
        "   • открыто онлайн и ментал хелп — оба формата",
    ]
    for i, line in enumerate(lines, start=2):
        help_ws.cell(row=i, column=1, value=line)
    help_ws.column_dimensions["A"].width = 78

    list_ws = wb.create_sheet("Значения")
    list_ws.sheet_state = "hidden"
    for i, val in enumerate(STATUS_VALUES, start=1):
        list_ws.cell(row=i, column=1, value=val)

    dv = DataValidation(
        type="list",
        formula1="=Значения!$A$1:$A$4",
        allow_blank=False,
        showDropDown=False,
    )
    dv.error = "Выберите значение из списка"
    dv.errorTitle = "Неверное значение"
    ws.add_data_validation(dv)
    dv.add(f"B{first_data_row}:H{last_data_row}")

    note = ws.cell(row=last_data_row + 2, column=1)
    note.value = (
        "Подсказка: заполните ячейки и отправьте файл боту. "
        "Даты в строке 1 можно менять каждую неделю."
    )
    note.font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=last_data_row + 2, start_column=1, end_row=last_data_row + 2, end_column=8)

    return wb


def main():
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "templates" / "schedule_week_template.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = build_template()
    wb.save(out)
    print(out)

    from schedule_excel import parse_schedule_xlsx

    try:
        rules = parse_schedule_xlsx(str(out))
        print(f"parse_ok slots={len(rules)} open={sum(1 for r in rules if r['is_open'])}")
    except ValueError as e:
        if "нет ни одного открытого" in str(e):
            print("parse_ok structure (all closed — fill before upload)")
        else:
            raise


if __name__ == "__main__":
    main()
