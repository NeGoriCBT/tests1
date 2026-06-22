#!/usr/bin/env python3
import sys
from openpyxl import load_workbook

path = sys.argv[1] if len(sys.argv) > 1 else "sample_schedule.xlsx"
wb = load_workbook(path, data_only=True)
for sn in wb.sheetnames:
    ws = wb[sn]
    print("SHEET:", sn, "rows:", ws.max_row, "cols:", ws.max_column)
    for r in range(1, min(ws.max_row + 1, 50)):
        row = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 25))]
        if any(x is not None and str(x).strip() for x in row):
            print(r, row)
    print("---")
wb.close()

try:
    from schedule_excel import parse_schedule_xlsx
    rules = parse_schedule_xlsx(path)
    print("PARSED:", len(rules), "rules")
    for x in rules:
        print(x)
except Exception as e:
    print("PARSE ERROR:", e)
