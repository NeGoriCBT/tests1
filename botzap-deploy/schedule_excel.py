"""
Импорт расписания из Excel.

Формат сетки (ваш файл):
  Строка 2: понедельник | вторник | ... | воскресенье
  Столбец A: 09:00, 10:00, ... 22:00
  Ячейки: закрыто | открыто онлайн | открыто ментал хелп

Старый формат (строки):
  День недели | Начало | Конец | Открыто
"""
from __future__ import annotations

import re
from datetime import date, time, datetime, timedelta
from typing import List, Dict, Any

from openpyxl import load_workbook

WEEKDAY_MAP = {
    "понедельник": 0, "пн": 0,
    "вторник": 1, "вт": 1,
    "среда": 2, "ср": 2,
    "четверг": 3, "чт": 3,
    "пятница": 4, "пт": 4,
    "суббота": 5, "сб": 5,
    "воскресенье": 6, "вс": 6,
}


def _norm(s) -> str:
    if s is None:
        return ""
    return str(s).strip().lower()


def _parse_time_val(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, time):
        return val.strftime("%H:%M")
    if isinstance(val, datetime):
        return val.strftime("%H:%M")
    s = str(val).strip().replace(".", ":")
    m = re.match(r"^(\d{1,2}):(\d{2})$", s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    return None


def _cell_open(val) -> tuple[bool, str]:
    """Открыт ли слот и подпись (онлайн / ментал хелп)."""
    s = _norm(val)
    if not s or s == "закрыто":
        return False, ""
    if "открыто" in s or s in ("да", "yes", "1", "+"):
        return True, s
    return False, s


def _find_weekday_header_row(rows: list) -> tuple[int, Dict[int, int]] | None:
    """Возвращает (номер_строки, {col_index: weekday 0-6})."""
    for i, row in enumerate(rows):
        col_map: Dict[int, int] = {}
        for j, cell in enumerate(row):
            wd = WEEKDAY_MAP.get(_norm(cell))
            if wd is not None:
                col_map[j] = wd
        if len(col_map) >= 3:
            return i, col_map
    return None


def _parse_grid(rows: list) -> List[Dict[str, Any]]:
    found = _find_weekday_header_row(rows)
    if not found:
        return []
    header_i, col_map = found
    slots: List[Dict[str, Any]] = []

    for row in rows[header_i + 1 :]:
        if not row:
            continue
        slot_time = None
        for cell in row:
            t = _parse_time_val(cell)
            if t:
                slot_time = t
                break
        if not slot_time:
            continue
        for col_i, weekday in col_map.items():
            if col_i >= len(row):
                continue
            is_open, label = _cell_open(row[col_i])
            slots.append({
                "weekday": weekday,
                "slot_time": slot_time,
                "is_open": is_open,
                "slot_label": label,
            })
    return slots


def _parse_rows_format(rows: list) -> List[Dict[str, Any]]:
    """Старый формат: день | начало | конец | открыто → почасовые слоты."""
    header_idx = 0
    for i, row in enumerate(rows[:10]):
        joined = " ".join(_norm(c) for c in row if c)
        if "день" in joined:
            header_idx = i
            break

    cols: Dict[str, int] = {}
    for j, cell in enumerate(rows[header_idx]):
        t = _norm(cell)
        if "день" in t:
            cols["day"] = j
        elif any(x in t for x in ("начало", "с ", "от", "from")):
            cols["from"] = j
        elif any(x in t for x in ("конец", "до", "to")):
            cols["to"] = j
        elif any(x in t for x in ("открыто", "статус")):
            cols["open"] = j

    if "day" not in cols or "from" not in cols or "to" not in cols:
        return []

    from datetime import timedelta as td

    def parse_hm(s):
        h, m = map(int, _parse_time_val(s).split(":"))
        return datetime(2000, 1, 1, h, m)

    slots = []
    for row in rows[header_idx + 1 :]:
        if not row or all(c is None for c in row):
            continue
        try:
            wd = WEEKDAY_MAP.get(_norm(row[cols["day"]]))
            if wd is None and _norm(row[cols["day"]]).isdigit():
                wd = int(_norm(row[cols["day"]]))
            if wd is None:
                continue
            is_open = _cell_open(row[cols["open"]])[0] if "open" in cols else True
            if not is_open:
                for hour in range(9, 23):
                    slots.append({"weekday": wd, "slot_time": f"{hour:02d}:00", "is_open": False, "slot_label": ""})
                continue
            start = parse_hm(row[cols["from"]])
            end = parse_hm(row[cols["to"]])
            cur = start
            while cur < end:
                slots.append({
                    "weekday": wd,
                    "slot_time": cur.strftime("%H:%M"),
                    "is_open": True,
                    "slot_label": "",
                })
                cur += td(hours=1)
        except (ValueError, IndexError, TypeError):
            continue
    return slots


def _find_date_header_row(rows: list) -> tuple[int, Dict[int, date]] | None:
    for i, row in enumerate(rows[:10]):
        col_dates: Dict[int, date] = {}
        for j, cell in enumerate(row or []):
            if j == 0:
                continue
            if isinstance(cell, datetime):
                col_dates[j] = cell.date()
            elif isinstance(cell, date):
                col_dates[j] = cell
        if len(col_dates) >= 3:
            return i, col_dates
    return None


def _weekday_header_row_index(rows: list, after_i: int) -> int:
    for i in range(after_i + 1, min(after_i + 4, len(rows))):
        row = rows[i] or []
        n = sum(1 for c in row if WEEKDAY_MAP.get(_norm(c)) is not None)
        if n >= 3:
            return i
    return after_i


def _parse_calendar_sheet(rows: list) -> List[Dict[str, Any]]:
    found = _find_date_header_row(rows)
    if not found:
        return []
    date_i, col_dates = found
    wd_i = _weekday_header_row_index(rows, date_i)
    slots: List[Dict[str, Any]] = []
    for row in rows[wd_i + 1 :]:
        if not row:
            continue
        slot_time = _parse_time_val(row[0] if len(row) > 0 else None)
        if not slot_time:
            continue
        for col_i, book_date in col_dates.items():
            if col_i >= len(row):
                continue
            is_open, label = _cell_open(row[col_i])
            slots.append({
                "book_date": book_date.isoformat(),
                "slot_time": slot_time,
                "is_open": is_open,
                "slot_label": label,
            })
    return slots


def _collect_dates_from_workbook(path: str) -> list[date]:
    wb = load_workbook(path, read_only=True, data_only=True)
    dates: list[date] = []
    for ws in wb.worksheets:
        if ws.title in ("Значения", "Справка"):
            continue
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            for cell in (row or [])[1:8]:
                if isinstance(cell, datetime):
                    dates.append(cell.date())
                elif isinstance(cell, date):
                    dates.append(cell)
    wb.close()
    return dates


def parse_schedule_calendar(path: str) -> tuple[List[Dict[str, Any]], str | None, str | None]:
    """Расписание по конкретным датам (листы «Неделя 1», «Неделя 2» …)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    slots: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        if ws.title in ("Значения", "Справка"):
            continue
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        slots.extend(_parse_calendar_sheet(rows))
    wb.close()
    if not slots:
        return [], None, None
    open_count = sum(1 for s in slots if s["is_open"])
    if open_count == 0:
        raise ValueError("В файле нет ни одного открытого слота")
    merged: Dict[tuple[str, str], Dict[str, Any]] = {}
    for s in slots:
        merged[(s["book_date"], s["slot_time"])] = s
    slots = list(merged.values())
    dates = _collect_dates_from_workbook(path)
    if not dates:
        dates = [date.fromisoformat(s["book_date"]) for s in slots]
    open_from = min(dates).isoformat()
    close_from = (max(dates) + timedelta(days=1)).isoformat()
    return slots, open_from, close_from


def import_schedule_from_xlsx(path: str) -> dict:
    """Календарное или недельное расписание."""
    cal_slots, open_from, close_from = parse_schedule_calendar(path)
    if cal_slots:
        return {
            "mode": "date",
            "slots": cal_slots,
            "open_from": open_from,
            "close_from": close_from,
        }
    rules = parse_schedule_xlsx(path)
    return {
        "mode": "weekday",
        "slots": rules,
        "open_from": None,
        "close_from": parse_schedule_close_from(path),
    }


def _hours(start: int, end: int) -> List[str]:
    """Часы start..end-1, например 9,13 → 09:00–12:00."""
    return [f"{h:02d}:00" for h in range(start, end)]


# Очно «ментал хелп» по календарю на картинке (июнь 2026 + 2–4 июля)
_IN_PERSON_JUNE_IMAGE: Dict[str, List[str]] = {
    "2026-06-04": _hours(9, 13),
    "2026-06-05": _hours(16, 20),
    "2026-06-06": _hours(10, 18),
    "2026-06-11": _hours(9, 13),
    "2026-06-12": _hours(16, 20),
    "2026-06-13": _hours(10, 18),
    "2026-06-18": _hours(9, 13),
    "2026-06-19": _hours(16, 20),
    "2026-06-21": _hours(10, 18),
    "2026-06-25": _hours(9, 13),
    "2026-06-28": _hours(10, 17),
    "2026-07-02": _hours(9, 13),
    "2026-07-03": _hours(16, 20),
    "2026-07-04": _hours(10, 18),
}


_ONLINE_MON_WED_TIMES = {
    0: ("10:00", "11:00"),  # пн
    1: ("16:00", "17:00", "18:00", "19:00"),  # вт
    2: ("10:00", "11:00"),  # ср
}


def _online_slots_for_june_2026() -> List[Dict[str, Any]]:
    """Все пн/вт/ср июня 2026 — шаблон из Excel."""
    import db

    out: List[Dict[str, Any]] = []
    for day in range(1, 31):
        d = date(2026, 6, day)
        times = _ONLINE_MON_WED_TIMES.get(d.weekday())
        if not times:
            continue
        for slot_time in times:
            out.append(
                {
                    "book_date": d.isoformat(),
                    "slot_time": slot_time,
                    "is_open": True,
                    "slot_label": db.SLOT_LABEL_ONLINE,
                }
            )
    return out


def apply_online_june_mon_tue_wed_preserving_bookings(
    *,
    open_from: str = "2026-06-01",
    close_from: str = "2026-07-06",
) -> dict:
    """
    Онлайн на все пн/вт/ср июня 2026 (10–11 и 16–19).
    Очные слоты и bookings не трогаем.
    """
    import db
    from db_conn import connect

    merged: Dict[tuple[str, str], Dict[str, Any]] = {}

    with connect() as c:
        for r in c.execute(
            """SELECT book_date, slot_time, slot_label FROM schedule_date_slots
               WHERE is_open=1 AND slot_label NOT LIKE '%онлайн%'"""
        ).fetchall():
            key = (r["book_date"], r["slot_time"])
            merged[key] = {
                "book_date": key[0],
                "slot_time": key[1],
                "is_open": True,
                "slot_label": r["slot_label"] or db.SLOT_LABEL_IN_PERSON,
            }

    for s in _online_slots_for_june_2026():
        key = (s["book_date"], s["slot_time"])
        merged[key] = s

    preserved: list[str] = []
    for b in db.get_active_booking_slots():
        key = (b["book_date"], b["book_time"])
        vt = b.get("visit_type") or "in_person"
        label = (
            db.SLOT_LABEL_ONLINE
            if vt == "online"
            else db.SLOT_LABEL_IN_PERSON
        )
        if key not in merged or not merged[key].get("is_open"):
            preserved.append(f"{key[0]} {key[1]}")
        merged[key] = {
            "book_date": key[0],
            "slot_time": key[1],
            "is_open": True,
            "slot_label": label,
        }

    online_dates = sorted({s["book_date"] for s in _online_slots_for_june_2026()})
    db.save_schedule_by_date(
        list(merged.values()),
        open_from=open_from,
        close_from=close_from,
    )
    return {
        "open_from": open_from,
        "close_from": close_from,
        "online_days": len(online_dates),
        "online_slots": len(_online_slots_for_june_2026()),
        "online_dates": online_dates,
        "preserved_bookings": preserved,
        "active_bookings": len(db.get_active_booking_slots()),
        "total_open_slots": sum(1 for s in merged.values() if s.get("is_open")),
    }


def apply_in_person_june_image_preserving_bookings() -> dict:
    """
    Очное расписание по календарю на картинке (5 недель).
    Онлайн-слоты из БД не меняются; bookings не трогаем.
    """
    import db
    from db_conn import connect

    merged: Dict[tuple[str, str], Dict[str, Any]] = {}

    with connect() as c:
        for r in c.execute(
            """SELECT book_date, slot_time, slot_label FROM schedule_date_slots
               WHERE is_open=1 AND slot_label LIKE '%онлайн%'"""
        ).fetchall():
            key = (r["book_date"], r["slot_time"])
            merged[key] = {
                "book_date": key[0],
                "slot_time": key[1],
                "is_open": True,
                "slot_label": r["slot_label"] or db.SLOT_LABEL_ONLINE,
            }

    for book_date, times in _IN_PERSON_JUNE_IMAGE.items():
        for slot_time in times:
            key = (book_date, slot_time)
            merged[key] = {
                "book_date": book_date,
                "slot_time": slot_time,
                "is_open": True,
                "slot_label": db.SLOT_LABEL_IN_PERSON,
            }

    preserved: list[str] = []
    for b in db.get_active_booking_slots():
        key = (b["book_date"], b["book_time"])
        vt = b.get("visit_type") or "in_person"
        label = (
            db.SLOT_LABEL_ONLINE
            if vt == "online"
            else db.SLOT_LABEL_IN_PERSON
        )
        if key not in merged or not merged[key].get("is_open"):
            preserved.append(f"{key[0]} {key[1]}")
        merged[key] = {
            "book_date": key[0],
            "slot_time": key[1],
            "is_open": True,
            "slot_label": label,
        }

    db.save_schedule_by_date(
        list(merged.values()),
        open_from="2026-06-01",
        close_from="2026-07-06",
    )
    in_person_dates = sorted(_IN_PERSON_JUNE_IMAGE.keys())
    return {
        "open_from": "2026-06-01",
        "close_from": "2026-07-06",
        "in_person_days": len(in_person_dates),
        "in_person_slots": sum(len(t) for t in _IN_PERSON_JUNE_IMAGE.values()),
        "preserved_bookings": preserved,
        "active_bookings": len(db.get_active_booking_slots()),
    }


def apply_schedule_preserving_bookings(path: str) -> dict:
    """
    Загрузить расписание из Excel, не трогая таблицу bookings.
    Слоты с активными записями остаются открытыми, даже если в файле закрыты.
    """
    import db

    imported = import_schedule_from_xlsx(path)
    if imported["mode"] != "date":
        raise ValueError("Нужен файл с датами (листы «Неделя 1», «Неделя 2» …)")

    merged: Dict[tuple[str, str], Dict[str, Any]] = {
        (s["book_date"], s["slot_time"]): dict(s) for s in imported["slots"]
    }
    preserved: list[tuple[str, str]] = []

    for b in db.get_active_booking_slots():
        key = (b["book_date"], b["book_time"])
        vt = b.get("visit_type") or "in_person"
        label = (
            db.SLOT_LABEL_ONLINE
            if vt == "online"
            else db.SLOT_LABEL_IN_PERSON
        )
        if key in merged:
            if not merged[key].get("is_open"):
                preserved.append(key)
            merged[key]["is_open"] = True
            if not merged[key].get("slot_label"):
                merged[key]["slot_label"] = label
        else:
            preserved.append(key)
            merged[key] = {
                "book_date": key[0],
                "slot_time": key[1],
                "is_open": True,
                "slot_label": label,
            }

    slots = list(merged.values())
    db.save_schedule_by_date(
        slots,
        open_from=imported["open_from"],
        close_from=imported["close_from"],
    )
    open_count = sum(1 for s in slots if s.get("is_open"))
    return {
        "open_from": imported["open_from"],
        "close_from": imported["close_from"],
        "open_slots": open_count,
        "preserved_bookings": [f"{d} {t}" for d, t in preserved],
        "active_bookings": len(db.get_active_booking_slots()),
    }


def parse_schedule_close_from(path: str) -> str | None:
    """Первый закрытый день = день после последней даты в Excel."""
    dates = _collect_dates_from_workbook(path)
    if not dates:
        return None
    return (max(dates) + timedelta(days=1)).isoformat()


def parse_schedule_valid_until(path: str) -> str | None:
    """Устар.: для совместимости."""
    close = parse_schedule_close_from(path)
    if not close:
        return None
    return (date.fromisoformat(close) - timedelta(days=1)).isoformat()


def parse_schedule_xlsx(path: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    slots = _parse_grid(rows)
    if not slots:
        slots = _parse_rows_format(rows)
    if not slots:
        raise ValueError(
            "Не удалось прочитать расписание. Нужна сетка (дни в строке, часы в столбце A) "
            "или таблица: День | Начало | Конец | Открыто"
        )
    open_count = sum(1 for s in slots if s["is_open"])
    if open_count == 0:
        raise ValueError("В файле нет ни одного открытого слота")
    return slots
